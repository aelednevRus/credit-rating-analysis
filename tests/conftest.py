"""Общие фикстуры для тестов credit_analysis."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from credit_analysis.extract import RawReport


def make_report(balance: dict[str, list[float]] | None = None,
                 pl: dict[str, list[float]] | None = None,
                 **kwargs) -> RawReport:
    """Собирает RawReport напрямую из словарей {ключ_показателя: [тек, пред, ...]}."""
    return RawReport(balance=balance or {}, pl=pl or {}, **kwargs)


@pytest.fixture
def report_factory():
    return make_report


def write_workbook(path: Path, info: dict, balance_rows: list[tuple],
                    pl_rows: list[tuple]) -> Path:
    """Строит минимальный .xlsx с тремя листами БФО по кодам РСБУ.

    balance_rows / pl_rows: список (наименование, код, [значения по периодам]).
    """
    wb = openpyxl.Workbook()

    ws_info = wb.active
    ws_info.title = "Сведения об организации"
    ws_info.append(["Наименование организации", info.get("name", "")])
    ws_info.append(["ИНН", info.get("inn", "")])

    ws_bal = wb.create_sheet("Бухгалтерский баланс")
    ws_bal.append(["Наименование показателя", "Код", "Тек.", "Пред.", "Позапр."])
    for name, code, vals in balance_rows:
        ws_bal.append([name, code, *vals])

    ws_pl = wb.create_sheet("Отчет о финансовых результатах")
    ws_pl.append(["Наименование показателя", "Код", "Тек.", "Пред."])
    for name, code, vals in pl_rows:
        ws_pl.append([name, code, *vals])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


@pytest.fixture
def workbook_factory(tmp_path):
    def _factory(name: str, info: dict, balance_rows: list[tuple],
                 pl_rows: list[tuple]) -> Path:
        return write_workbook(tmp_path / name, info, balance_rows, pl_rows)
    return _factory

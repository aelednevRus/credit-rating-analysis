"""Чтение бухгалтерской отчётности (БФО, РСБУ) из .xlsx.

Ожидается выгрузка с тремя листами:
    «Сведения об организации», «Бухгалтерский баланс»,
    «Отчет о финансовых результатах».

Из баланса берутся 3 отчётные даты (текущий / предыдущий / позапрошлый год),
из ОФР — 2 периода (текущий / предыдущий). Отрицательные значения в круглых
скобках и пропуски «-» нормализуются функцией `_to_number`.

Модуль устойчив к вариациям структуры файла: строки ищутся по коду строки
РСБУ (столбец с кодами 1200, 1210, ...), а не по фиксированным координатам.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# --- Коды строк РСБУ, используемые методикой -------------------------------

BALANCE_CODES = {
    "current_assets": "1200",      # Оборотные активы
    "inventory": "1210",           # Запасы
    "vat": "1220",                 # НДС по приобретённым ценностям
    "cash": "1250",                # Денежные средства
    "short_investments": "1240",   # Краткосрочные финансовые вложения
    "long_investments": "1170",    # Финансовые вложения (внеоборотные) — для холдинга
    "short_liabilities": "1500",   # Краткосрочные обязательства
    "equity": "1300",              # Капитал и резервы (СК)
    "total_assets": "1600",        # Активы, всего
    "long_liabilities": "1400",    # Долгосрочные обязательства
    "long_debt": "1410",           # Долгосрочные заёмные средства
    "short_debt": "1510",          # Краткосрочные заёмные средства
    "noncurrent_assets": "1100",   # Внеоборотные активы — для холдинга
}

PL_CODES = {
    "revenue": "2110",       # Выручка
    "sales_profit": "2200",  # Прибыль от продаж
    "interest_paid": "2330", # Проценты к уплате
    "net_profit": "2400",    # Чистая прибыль
    "other_income": "2340",  # Прочие доходы — для оценки качества прибыли
}

SHEET_INFO = "Сведения об организации"
SHEET_BALANCE = "Бухгалтерский баланс"
SHEET_PL = "Отчет о финансовых результатах"


@dataclass
class RawReport:
    """Нормализованные показатели отчётности одной компании."""

    company: str = "Без названия"
    inn: str = ""
    # Баланс: индекс 0 = текущий год, 1 = предыдущий, 2 = позапрошлый
    balance: dict[str, list[float]] = field(default_factory=dict)
    # ОФР: индекс 0 = текущий период, 1 = предыдущий
    pl: dict[str, list[float]] = field(default_factory=dict)
    source_file: str = ""

    def b(self, key: str, period: int = 0) -> float:
        """Значение баланса по ключу и периоду (0=тек., 1=пред., 2=позапр.)."""
        vals = self.balance.get(key, [])
        return vals[period] if period < len(vals) else 0.0

    def p(self, key: str, period: int = 0) -> float:
        """Значение ОФР по ключу и периоду (0=тек., 1=пред.)."""
        vals = self.pl.get(key, [])
        return vals[period] if period < len(vals) else 0.0


def _to_number(value) -> float:
    """Приводит ячейку отчётности к float.

    Обрабатывает: None и «-» -> 0; «(123)» -> -123; пробелы-разделители
    тысяч; запятую как десятичный разделитель.
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if s in {"", "-", "—", "–", "n/a", "N/A"}:
        return 0.0

    negative = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    s = s.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    # Уберём всё, кроме цифр, точки и знака минус
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in {"", "-", ".", "-."}:
        return 0.0
    try:
        num = float(s)
    except ValueError:
        return 0.0
    return -num if negative else num


def _find_code_column(rows: list[tuple]) -> int:
    """Определяет индекс столбца, где находятся коды строк РСБУ (1200, ...)."""
    code_pattern = re.compile(r"^\s*1[1-9]\d0\s*$|^\s*2[1-4]\d0\s*$")
    counts: dict[int, int] = {}
    for row in rows:
        for idx, cell in enumerate(row):
            if cell is None:
                continue
            if code_pattern.match(str(cell)):
                counts[idx] = counts.get(idx, 0) + 1
    if not counts:
        return -1
    return max(counts, key=counts.get)


def _is_number_like(cell) -> bool:
    """True, если ячейка содержит осмысленное число (не пустой разделитель)."""
    if cell is None:
        return False
    if isinstance(cell, (int, float)):
        return True
    s = str(cell).strip()
    if s in {"", "-", "—", "–"}:
        return False
    # Есть хотя бы одна цифра
    return any(ch.isdigit() for ch in s)


def _numeric_columns(row: tuple, code_col: int) -> list[float]:
    """Возвращает числовые значения строки правее столбца с кодом.

    Официальные выгрузки ГИР БО содержат пустые колонки-разделители между
    периодами (merged-ячейки). Мы оставляем только колонки с реальными
    числами, поэтому индексы периодов идут подряд: 0 = текущий год,
    1 = предыдущий, 2 = позапрошлый.
    """
    values = []
    for idx, cell in enumerate(row):
        if idx <= code_col:
            continue
        if _is_number_like(cell):
            values.append(_to_number(cell))
    return values


def _read_sheet_by_codes(ws, code_map: dict[str, str]) -> dict[str, list[float]]:
    """Читает лист, сопоставляя коды строк с ключами показателей."""
    rows = list(ws.iter_rows(values_only=True))
    code_col = _find_code_column(rows)
    if code_col < 0:
        return {}

    inverse = {code: key for key, code in code_map.items()}
    result: dict[str, list[float]] = {}
    for row in rows:
        if code_col >= len(row) or row[code_col] is None:
            continue
        code = str(row[code_col]).strip()
        key = inverse.get(code)
        if key is None:
            continue
        result[key] = _numeric_columns(row, code_col)
    return result


def _read_company_info(wb) -> tuple[str, str]:
    """Извлекает наименование и ИНН со ли��та «Сведения об организации»."""
    name, inn = "Без названия", ""
    if SHEET_INFO not in wb.sheetnames:
        return name, inn
    ws = wb[SHEET_INFO]
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() for c in row if c is not None]
        line = " ".join(cells)
        low = line.lower()
        if ("наименование" in low or "организац" in low) and name == "Без названия":
            # Значение обычно в соседней непустой ячейке
            for c in reversed(cells):
                if "наименован" not in c.lower() and "организац" not in c.lower():
                    name = c
                    break
        if "инн" in low and not inn:
            m = re.search(r"\b(\d{10,12})\b", line)
            if m:
                inn = m.group(1)
    return name, inn


def _pick_sheet(wb, target: str) -> "openpyxl.worksheet.worksheet.Worksheet | None":
    """Находит лист по точному или частичному совпадению имени."""
    if target in wb.sheetnames:
        return wb[target]
    key = target.lower().split()[0]
    for name in wb.sheetnames:
        if key in name.lower():
            return wb[name]
    return None


def extract_report(path: str | Path) -> RawReport:
    """Читает один .xlsx-файл БФО и возвращает нормализованные показатели."""
    path = Path(path)
    # read_only=True (потоковый парсер) на части реальных выгрузок ГИР БО
    # молча отдаёт пустой лист (нет <dimension> или нестандартный XML) —
    # грузим лист целиком, отчётности БФО небольшие, это не проблема памяти.
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    name, inn = _read_company_info(wb)
    report = RawReport(company=name, inn=inn, source_file=path.name)

    ws_bal = _pick_sheet(wb, SHEET_BALANCE)
    if ws_bal is not None:
        report.balance = _read_sheet_by_codes(ws_bal, BALANCE_CODES)

    ws_pl = _pick_sheet(wb, SHEET_PL)
    if ws_pl is not None:
        report.pl = _read_sheet_by_codes(ws_pl, PL_CODES)

    wb.close()

    # Fallback имени: если в «Сведениях» не нашли — берём имя файла
    if report.company == "Без названия":
        report.company = path.stem

    return report


if __name__ == "__main__":
    import sys

    for arg in sys.argv[1:]:
        rep = extract_report(arg)
        print(f"{rep.company} (ИНН {rep.inn or '—'}) из {rep.source_file}")
        print("  Баланс:", rep.balance)
        print("  ОФР:   ", rep.pl)
